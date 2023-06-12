import openpyxl

from .resource import get_alias_excel


class AliasManager:
    _alias_id_map: dict[str, int] = {}

    def __init__(self):
        self.load_from_excel()

    @staticmethod
    def clean_alias(alias: str) -> str:
        return alias.removesuffix('（')

    def load_from_excel(self):
        wb = openpyxl.load_workbook(get_alias_excel())
        sheet = wb['工作表1']

        for row in range(2, sheet.max_row + 1):  # 从第二行开始
            song_id: str = sheet.cell(row=row, column=1).value  # A 列存放歌曲 ID
            aliases: str = sheet.cell(row=row, column=3).value  # C 列存放别名列表

            if song_id is not None and aliases is not None:
                for alias in aliases.split(','):
                    self._alias_id_map[self.clean_alias(alias)] = int(song_id)

    def get_song_id(self, alias: str) -> int:
        return self._alias_id_map.get(self.clean_alias(alias), 0)
