import os.path
import re
import time
from datetime import date
from typing import Optional, Any, Union

import aiohttp
import matplotlib.pyplot as plt
import nonebot
from aiohttp_retry import ExponentialRetry, RetryClient
from matplotlib.font_manager import FontProperties
from nonebot import on_command, require
from nonebot.adapters import Bot
from nonebot.adapters.cqhttp import MessageSegment, MessageEvent
from nonebot.permission import SUPERUSER
from openpyxl import Workbook
from tenacity import retry, stop_after_attempt

from utils.mb2pkg_mokalogger import getlog
from utils.mb2pkg_public_plugin import datediff, get_time, now_datetime
from utils.mb2pkg_text2pic import draw_image
from .config import Config

scheduler = require('nonebot_plugin_apscheduler').scheduler

match_bandori_track = on_command('分数线', priority=5)
match_bandori_event_list = on_command('活动列表', priority=5)
match_make_chart_excel = on_command('生成歌曲列表', priority=5, permission=SUPERUSER)

log = getlog()

temp_absdir = nonebot.get_driver().config.temp_absdir
FONTPATH = os.path.join(Config().font_absdir, 'NotoSansMonoCJKsc-Regular.otf')

# TODO 对需要反复上bestdori查询的内容使用缓存


@match_bandori_track.handle()
async def bandori_track_handle(bot: Bot, event: MessageEvent):
    info = re.match(r'^(\d{1,3})(JP|EN|TW|CN|KR)(50|100|300|500|1000|2000|5000|10000)$', str(event.get_message()).strip())
    evt = info.group(1)
    server = info.group(2)
    rank = info.group(3)

    try:
        list_track_image_path = await list_track(int(evt), server, rank)
        msg = MessageSegment.image(file=f'file:///{list_track_image_path}')
        await bot.send(event, msg)
        event_prediction_image_path = await event_prediction(int(evt), server, rank)
        msg = MessageSegment.image(file=f'file:///{event_prediction_image_path}')
        await bot.send(event, msg)
    except IndexError:
        msg = str(event.get_message()).strip() + '的档线尚未在bestdori记录'
        await bot.send(event, msg)


@match_bandori_event_list.handle()
async def bandori_event_list_handle(bot: Bot, event: MessageEvent):
    server = str(event.get_message()).strip() or 'JP'
    list_event_image_path = await list_event(server)
    msg = MessageSegment.image(file=f'file:///{list_event_image_path}')

    await bot.send(event, msg)


@match_make_chart_excel.handle()
async def make_chart_excel_handle(bot: Bot, event: MessageEvent):
    try:
        await bot.send(event, '正在生成，请等待...')
        await make_chart_excel()
        msg = '已生成完毕！'
    except Exception as e:
        msg = f'生成时发生错误{e}，请前往服务端查看日志'
        log.exception(e)

    await bot.send(event, msg)


@retry(stop=stop_after_attempt(5))
async def get_event_track(event: Union[str, int], server: str, rank: Union[str, int]) -> list[dict[str, int]]:
    """
    返回包含活动pt观测值的列表，每个元素是一个字典，time关键字(int)为unix时间戳，ep关键字(int)为pt观测值

    :param event: bestdori活动序号(0~999)
    :param server: bestdori服务器名称 JP, EN, TW, CN, KR
    :param rank: bestdori活动排名 100, 500, 1000, 2000, 5000, 10000
    :return: 返回包含活动pt观测值的列表
    """

    server_dict = {'JP': 0, 'EN': 1, 'TW': 2, 'CN': 3, 'KR': 4}
    async with aiohttp.ClientSession() as session:
        url = f'https://bestdori.com/api/tracker/data?server={server_dict[server]}&event={event}&tier={rank}'
        async with session.get(url=url, timeout=5) as r:
            # 从网页获取json
            event_rank_data = (await r.json())['cutoffs']
            log.debug('从bestdori获取到了pt观测值json')
            log.debug(event_rank_data)

    return event_rank_data


@retry(stop=stop_after_attempt(5))
async def get_song_list() -> dict[str, Any]:
    """
    返回bestdori的歌曲meta列表

    :return: 返回歌曲meta列表
    """

    async with aiohttp.ClientSession() as session:
        url = 'https://bestdori.com/api/songs/all.7.json'
        async with session.get(url=url, timeout=5) as r:
            # 从网页获取json
            song_list_data = (await r.json())
            log.debug('从bestdori获取到了歌曲meta列表')

    return song_list_data


@retry(stop=stop_after_attempt(5))
async def get_band_list() -> dict[str, dict[str, list[str]]]:
    """
    返回bestdori的乐队列表

    :return: 返回乐队列表
    """

    async with aiohttp.ClientSession() as session:
        url = 'https://bestdori.com/api/bands/all.1.json'
        async with session.get(url=url, timeout=5) as r:
            # 从网页获取json
            band_list_data = (await r.json())
            log.debug('从bestdori获取到了乐队列表')

    return band_list_data


@retry(stop=stop_after_attempt(5))
async def get_event_archives() -> dict[str, Any]:
    """
    返回包含至今为止所有的活动基本数据（ID、类型、起止时间、名称）

    :return: 以字典形式返回至今为止所有的活动基本数据
    """

    async with aiohttp.ClientSession() as session:
        async with session.get(url='https://bestdori.com/api/events/all.6.json', timeout=5) as r:
            # 从网页获取json
            r.encoding = 'utf-8'
            event_archives = await r.json()
            log.debug('从bestdori获取到了活动基本数据')
            log.debug(event_archives)

    return event_archives


async def list_event(server: str) -> str:
    """
    列出指定服务器上，至今为止的活动列表

    :param server: 服务器名称
    :return: 返回生成图片路径
    """

    result = []
    event_type_dict = {'story': '一般', 'challenge': 'ＣＰ', 'versus': '对邦', 'live_try': 'ＥＸ', 'mission_live': '任务', 'festival': '队邦', 'medley': '三连'}
    server_dict = {'JP': 0, 'EN': 1, 'TW': 2, 'CN': 3, 'KR': 4}

    # 从bestdori获取数据
    event_archives = await get_event_archives()

    # i从1开始的原因:event_archives的起点是1而不是0
    i = 1
    while str(i) in event_archives and event_archives[str(i)]['startAt'][server_dict[server]] is not None:
        j = event_archives[str(i)]
        event_id = '{:<3d}'.format(i)  # 活动编号，从1开始，左对齐，宽度3，不足补空格
        event_type = event_type_dict[j['eventType']]  # 活动类型，见上字典
        event_start = get_time("%Y-%m-%d %H:%M:%S", j['startAt'][server_dict[server]][:-3])  # 活动开始时间
        event_end = get_time("%Y-%m-%d %H:%M:%S", j['endAt'][server_dict[server]][:-3])  # 活动结束时间
        event_name = j['eventName'][server_dict[server]]  # 活动名称
        result.append(' '.join([event_id, event_type, event_name]))
        result.append(f'  {event_start}～{event_end}')
        i += 1
    log.debug(f'已生成活动列表，活动总数：{i - 1}')

    head = [f'活动列表 来源：bestdori.com', f'制图时间：{now_datetime()}', '']  # 添加描述头

    savepath = os.path.join(temp_absdir, f'event_list_{server}.jpg')
    await draw_image(head + result, savepath)

    return savepath


async def list_track(event: int, server: str, rank: Union[str, int]) -> str:
    """
    活动追踪器，生成一个活动pt观测值及其变化速率的表

    :param event: bestdori活动序号(0~999)
    :param server: bestdori服务器名称 JP, EN, TW, CN, KR
    :param rank: bestdori活动排名 100, 1000, 2000
    :return: 返回生成图片路径
    """

    result = []
    event_type_dict = {'story': '一般', 'challenge': 'ＣＰ', 'versus': '对邦', 'live_try': 'ＥＸ', 'mission_live': '任务', 'festival': '队邦', 'medley': '三连'}
    server_dict = {'JP': 0, 'EN': 1, 'TW': 2, 'CN': 3, 'KR': 4}

    event_archives = await get_event_archives()
    # event为0时由moka帮你查询最新的活动序号
    if event == 0:
        i = 1
        while str(i) in event_archives and event_archives[str(i)]['startAt'][server_dict[server]] is not None:
            i += 1
        event = i - 1
        log.warn(f'event为0，查询最新一期，最新一期为{str(i - 1)}')

    # 从bestdori获取数据
    event_archives = event_archives[str(event)]
    event_track = await get_event_track(event, server, rank)

    event_type = event_type_dict[event_archives['eventType']]  # 活动类型，见上字典
    event_start = get_time("%Y-%m-%d %H:%M:%S", event_archives['startAt'][server_dict[server]][:-3])  # 活动开始时间
    event_end = get_time("%Y-%m-%d %H:%M:%S", event_archives['endAt'][server_dict[server]][:-3])  # 活动结束时间
    event_name = event_archives['eventName'][server_dict[server]]  # 活动名称
    log.debug(f'活动基本信息解析完成，{event_type} {event_start} {event_end} {event_name}')

    # 计算观测点密度
    observe_num = len(event_track)
    observe_last = (event_track[-1]['time'] / 1000 - int(event_archives['startAt'][server_dict[server]][:-3])) / 3600
    observe_density = observe_num / observe_last

    # 添加描述头
    head = [f'{event} {server}{rank} {event_type} {event_name}',
            f'起止时间：{event_start} ～ {event_end}',
            '数据来源：bestdori活动PT&排名追踪器 (bestdori.com)',
            f'制图时间：{now_datetime()}',
            '',
            '观测点数量：%d 已开始时间：%.1fhr 观测点密度：%.2f个/hr' % (observe_num, observe_last, observe_density),
            f'时效性：{datediff(time.time(), event_track[-1]["time"] / 1000)}',
            '',
            '样本    时间         已开始      观测值       增长          增速',
            '==================================================================']

    # 初始化数组 real_time:时间 past_time:已开始 event_pt:观测值 dx:增长 spd:增速  元素类型全部为int
    real_time, past_time, event_pt, dx, spd = [], [], [], [], []

    # 导入t和x的数据，计算pst的数据，t单位为ms，pst单位为hr，元素类型全部为int
    for i in event_track:
        real_time.append(i['time'])
        event_pt.append(i['ep'])
        past_time.append((i['time'] - int(event_archives['startAt'][server_dict[server]])) / 3600000)

    # dx和spd 的第一个数据需要单独计算
    dx.append(event_track[0]['ep'])
    spd.append(dx[0] / past_time[0])

    # 计算dx的值，int，计算spd的值，单位pt/hr，int
    for i in range(len(real_time) - 1):
        dx.append(event_pt[i + 1] - event_pt[i])
        try:
            spd.append(dx[i + 1] / (past_time[i + 1] - past_time[i]))
        except ZeroDivisionError:  # 偶尔出现同时提交导致bestdori会出现两个相同时间的数据，导致除数为0
            spd.append(0)

    # 输出数据到result
    for i in range(len(real_time)):
        result.append('{:>3d}  {:>11s}   {:>5.1f}hr   {:>8d}pt   {:>7d}pt   {:>7d}pt/hr'.format(
            i + 1, get_time("%m-%d %H:%M", real_time[i] / 1000), past_time[i], event_pt[i], dx[i], int(spd[i])
        ))
    log.debug(f'已计算完所有的数据，观测点数量：{len(result)}')

    savepath = os.path.join(temp_absdir, f'event_track_{event}{server}{rank}.jpg')
    await draw_image(head + result, savepath)

    return savepath


async def event_prediction(event: int, server: str, rank: Union[str, int]) -> str:
    """
    活动预测器，生成一个活动目前的pt预测值的折线图

    :param event: bestdori活动序号(0~999)
    :param server: bestdori服务器名称 JP, EN, TW, CN, KR
    :param rank: bestdori活动排名 100, 1000, 2000
    :return: 返回生成图片路径
    """

    def normalization(d: list, t_start: int) -> dict[str, list]:
        """归一化时间和分数序列"""
        # x0和y0是原始列表，赋一个0作为起点值
        x0 = [0] + [k['time'] - t_start for k in d]
        y0 = [0] + [k['ep'] for k in d]
        x_max = max(x0)
        y_max = max(y0)
        # x和y是归一化后的值
        x = [k / x_max for k in x0]
        y = [k / y_max for k in y0]
        return {'x': x, 'y': y}

    server_dict = {'JP': 0, 'EN': 1, 'TW': 2, 'CN': 3, 'KR': 4}
    result = []

    event_archives = await get_event_archives()
    # event为0时由moka帮你查询最新的活动序号
    if event == 0:
        i = 1
        while str(i) in event_archives and event_archives[str(i)]['startAt'][server_dict[server]] is not None:
            i += 1
        event = i - 1
        log.warn(f'event为0，查询最新一期，最新一期为{str(i - 1)}')

    # 判断活动类型
    event_archives = list(event_archives.values())
    event_type = event_archives[event - 1]['eventType']
    event_name = event_archives[event - 1]['eventName'][server_dict[server]]

    log.debug(f'将会对{event}{server}{rank}进行档线追踪与预测')
    log.debug(f'活动类型是{event_type}')

    # 找到同类型的最近5期有效活动数据
    history_track = []
    i = len(event_archives)
    while i > 0 and len(history_track) < 5:
        i -= 1
        ea = event_archives[i]
        # 如果查的是国服活动，这里就会跳过那些日服开了而国服还没开的活动，而不用依靠空列表检测浪费时间
        # noinspection PyTypeChecker
        if ea['endAt'][server_dict[server]] is None:
            continue
        # 首先判断类型是否相同
        # noinspection PyTypeChecker
        if ea['eventType'] == event_type:
            t = await get_event_track(i + 1, server, rank)
            if not t:  # 当bestdori完全没有统计这期数据时，返回的是个空列表
                log.warn(f'bestdori的{i + 1}{server}{rank}统计数据为空，跳过')
                continue
            else:
                # noinspection PyTypeChecker
                t_end = int(event_archives[i]['endAt'][server_dict[server]]) + 1000  # 注意时间单位为ms
                if t_end - t[-1]['time'] > 0:  # 当bestdori没有将这期数据统计完时，最后一个的时间戳和结活时间戳不相等
                    log.warn(f'bestdori的{i + 1}{server}{rank}统计数据不完全，跳过')
                    continue
                else:
                    log.debug(f'bestdori的{i + 1}{server}{rank}统计数据有效')
                    # noinspection PyTypeChecker
                    history_track.append(normalization(t, int(event_archives[i]['startAt'][server_dict[server]])))

    log.debug(f'找到了{len(history_track)}期有效活动数据')

    event_point_list = await get_event_track(event, server, rank)
    event_time_start = int(event_archives[event - 1]['startAt'][server_dict[server]])
    event_time_end = int(event_archives[event - 1]['endAt'][server_dict[server]]) + 1000

    # 计算每个点的预测档线
    for single_point in event_point_list:
        point_time_nm = (single_point['time'] - event_time_start) / (event_time_end - event_time_start)
        point_ep = single_point['ep']

        # 计算该时间所对应的期望分数百分比，方法为在每条折线上计算其投影，并求平均值
        point_ep_nm = 0
        for item in history_track:
            for i in range(len(item['x']) - 1):
                x1 = item['x'][i]
                x2 = item['x'][i + 1]
                y1 = item['y'][i]
                y2 = item['y'][i + 1]
                # 一般情况，时间处在两点之间
                if x1 < point_time_nm < x2:
                    point_ep_nm += (y2 - y1) * (point_time_nm - x1) / (x2 - x1) + y1  # 两点式直线方程
                    break
                # 特殊情况，时间处在两点的端点上，例如开始和结束
                elif point_time_nm == x1:
                    point_ep_nm += y1
                    break
                elif point_time_nm == x2:
                    point_ep_nm += y2
                    break
        # 求出了该时间所对应的期望分数百分比所对应的5个投影，对着5个投影的值求平均值
        point_ep_nm /= len(history_track)
        # 有个时候bestdori的end时间写错了几秒钟导致最后一个分数点的归一值为0
        if point_ep_nm == 0:
            continue
        result.append({'time': single_point['time'], 'ep': int(point_ep / point_ep_nm)})

    log.debug(f'所有点的预测线计算完成，理论计算{len(event_point_list)}个点，实际计算{len(result)}个点')

    # 导出数据到数组，便于作图，时间单位为已经过小时，分数单位为万
    real_point_x = []
    real_point_y = []
    prediction_point_x = []
    prediction_point_y = []
    # 重置单位，时间单位为小时，分数单位为万
    for i in range(len(result)):
        real_point_x.append((event_point_list[i]['time'] - event_time_start) / 3600000)
        real_point_y.append(event_point_list[i]['ep'] / 10000)
        prediction_point_x.append((result[i]['time'] - event_time_start) / 3600000)
        prediction_point_y.append(result[i]['ep'] / 10000)

    # 开始制图
    font = FontProperties(fname=FONTPATH, size=14)
    plt.suptitle(f'{event}期 {server}{rank}位活动预测线 ({datediff(time.time(), event_point_list[-1]["time"] / 1000)})\n{event_name}',
                 fontproperties=font)  # 标题
    plt.plot(prediction_point_x, prediction_point_y, 'red', lw=2, label='预测', )  # 预测变量
    plt.plot(real_point_x, real_point_y, 'green', lw=2, label='实时', marker='.')  # 实时档线变量
    font = FontProperties(fname=FONTPATH, size=11)
    plt.xlim(-10, 10 + (event_time_end - event_time_start) / 3600000)  # x轴范围
    plt.xlabel('已开始时间（小时）', fontproperties=font)  # x轴标签
    plt.ylabel('分数（万）', fontproperties=font)  # y轴标签
    plt.axvline(x=(event_time_end - event_time_start) / 3600000, ls=":", c="blue")  # 添加垂线，垂足在结活时间
    plt.annotate(int(prediction_point_y[-1] * 10000),
                 xy=(prediction_point_x[-1], prediction_point_y[-1]), xycoords='data',
                 xytext=(-40, +20), textcoords='offset points', fontsize=10,
                 arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=.3"))  # 指向数据点的箭头
    plt.annotate(int(real_point_y[-1] * 10000),
                 xy=(real_point_x[-1], real_point_y[-1]), xycoords='data',
                 xytext=(0, -20), textcoords='offset points', fontsize=10,
                 arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=.3"))  # 指向数据点的箭头
    plt.legend(loc='best', prop=font)  # 图例位置
    plt.grid()  # 网格

    # 生成图片
    savepath = os.path.join(temp_absdir, f'event_prediction_{event}{server}{rank}.jpg')
    plt.savefig(savepath)
    log.debug(f'已生成图片，保存在{savepath}')
    plt.clf()

    return savepath


class BandoriSongMeta:  # 原本打算用pydantic，但是发现bestdori居然把"1"、"2"等作为json的key，导致model类的属性不符合py变量命名规则

    def __init__(self,
                 data: dict,
                 song_id: int):
        self.data = data
        self.song_id = song_id
        self.special = len(self.data['difficulty']) == 5

    def __getitem__(self, key):
        return self.data[key]


class BandoriChartInfo:

    def __init__(self,
                 bandori_song_meta: BandoriSongMeta,
                 band_map: dict,
                 difficulty: int):
        self.mt = bandori_song_meta
        self.bm = band_map
        self.df = difficulty  # 注意：bestdori是用字符"1"来表示难度的，而不是数字1，因此需要大量str(self.df)

        self.id: int = self.mt.song_id
        self.chart_id: int = self._get_chart_id()  # 谱面编号，区别于歌曲编号，sp谱面的编号是id+10000，ex谱面的编号等于id
        self.band: str = self._get_band_name()
        self.tag: str = self._get_song_type()  # 乐曲类型
        self.bpm: str = self._get_real_bpm()  # bpm值或范围）
        self.release_jp: Optional[date] = self._get_release(0)
        self.release_cn: Optional[date] = self._get_release(3)
        self.note: int = self._get_note()  # 物量
        self.title: str = self._get_title()
        self.level: int = self._get_level()  # 难度（数字，例如25 26）
        self.level_class: str = self._get_class()  # 难度（类型，例如ex sp）
        self.lyricist: str = self._get_lyricist()
        self.composer: str = self._get_composer()
        self.arranger: str = self._get_arranger()

    def _get_chart_id(self) -> int:
        return self.mt.song_id + 10000 if self.df == 4 else self.mt.song_id

    def _get_band_name(self) -> str:
        return self.bm[str(self.mt['bandId'])]['bandName'][0]

    def _get_song_type(self) -> str:
        return {
            'normal': 'Original',  # 原唱
            'anime': 'Cover',  # 翻唱
            'tie_up': 'Extra',  # 联动
        }.get(self.mt['tag'], 'Unknown')

    def _get_real_bpm(self) -> str:
        bpm_list = [_['bpm'] for _ in self.mt['bpm'][str(self.df)]]
        max_bpm = max(bpm_list)
        min_bpm = min(bpm_list)
        if max_bpm == min_bpm:
            result = str(max_bpm)
        else:
            result = f'{min_bpm}-{max_bpm}'
        return result

    def _get_release(self, server: int) -> Optional[date]:
        # server: 0-JP, 1-EN, 2-TW, 3-CN, 4-KR
        if self.df == 4:  # Special难度
            if self.id == 359:  # TODO HELL! or HELL? ，bestdori没有标记其sp发布时间
                if server == 0:
                    result = date(2021, 9, 19)
                else:
                    result = None  # TODO 因为国服没出啦
            else:
                result = self._get_time(self.mt['difficulty']['4']['publishedAt'][server])
        else:
            result = self._get_time(self.mt['publishedAt'][server])
        return result

    def _get_note(self) -> int:
        return self.mt['notes'][str(self.df)]

    def _get_title(self) -> str:
        return self.mt['musicTitle'][0]

    def _get_level(self) -> int:
        return self.mt['difficulty'][str(self.df)]['playLevel']

    def _get_class(self) -> str:
        return {
            0: 'Easy',
            1: 'Normal',
            2: 'Hard',
            3: 'Expert',
            4: 'Special',
        }.get(self.df, 'Unknown')

    def _get_lyricist(self) -> str:
        return self.mt['lyricist'][0]

    def _get_composer(self) -> str:
        return self.mt['composer'][0]

    def _get_arranger(self) -> str:
        return self.mt['arranger'][0]

    @staticmethod
    def _get_time(timestamp: str) -> Optional[date]:
        try:
            return date.fromtimestamp(int(timestamp[:-3]))
        except TypeError:
            pass  # 等价于 return None


@scheduler.scheduled_job('cron', hour='4')
async def make_chart_excel() -> None:

    # 一个原曲，两个国际服特供曲
    exclude_song_id = [273, 1000, 1001]
    headline = [
        'Id',
        'Band',
        'Tag',
        'BPM',
        'Release(JP)',
        'Release(CN)',
        'Note',
        'Title',
        'Level',
        'Class',
        # 'Lyricist',
        # 'Composer',
        # 'Arranger',
    ]
    # 对应BandoriChartInfo类中的属性名，将使用__getattribute__来获取属性
    column_to_write = {
        1: 'chart_id',
        2: 'band',
        3: 'tag',
        4: 'bpm',
        5: 'release_jp',
        6: 'release_cn',
        7: 'note',
        8: 'title',
        9: 'level',
        10: 'level_class',
        # 11: 'lyricist',
        # 12: 'composer',
        # 13: 'arranger',
    }

    start = time.time()

    book = Workbook()
    sheet_all = book.create_sheet(title='charts_all', index=0)
    sheet_ex = book.create_sheet(title='charts_ex', index=1)
    sheet_sp = book.create_sheet(title='charts_sp', index=2)

    # 写标题行
    for col, item in enumerate(headline, start=1):
        sheet_all.cell(column=col, row=1, value=item)
        sheet_ex.cell(column=col, row=1, value=item)
        sheet_sp.cell(column=col, row=1, value=item)

    all_chart_list: list[BandoriChartInfo] = []
    ex_chart_list: list[BandoriChartInfo] = []
    sp_chart_list: list[BandoriChartInfo] = []

    # 从bestdori爬取所需信息
    retry_options = ExponentialRetry(attempts=5)
    async with RetryClient(retry_options=retry_options) as client:
        # 获取bestdori的歌曲meta列表，仅包含歌曲id，不包含其他任何信息，其json为 {"1":{},"2":{},"3":{}, ... }
        async with client.get(url='https://bestdori.com/api/songs/all.0.json', timeout=5) as response:
            simple_song_list = await response.json()
            log.info('已获取到bestdori的歌曲meta列表')
        # 获取bestdori的乐队名称
        async with client.get(url='https://bestdori.com/api/bands/all.1.json', timeout=5) as response:
            band_list = await response.json()
            log.info('已获取到bestdori的乐队名称')
        # 下载歌曲meta并格式化为谱面
        for song_id in simple_song_list:  # song_id:: str
            song_id: int = int(song_id)
            if song_id in exclude_song_id:
                continue
            async with client.get(url=f'https://bestdori.com/api/songs/{song_id}.json', timeout=5) as response:
                song_meta = BandoriSongMeta(await response.json(), song_id)
                # 将歌曲meta（BandoriSongMeta）格式化为谱面（BandoriChartInfo）
                ex_info = BandoriChartInfo(song_meta, band_list, 3)
                ex_chart_list.append(ex_info)
                all_chart_list.append(ex_info)
                log.info(f'歌曲{ex_info.title}，谱面id={ex_info.chart_id}的信息已获取')
                if song_meta.special:
                    sp_info = BandoriChartInfo(song_meta, band_list, 4)
                    sp_chart_list.append(sp_info)
                    all_chart_list.append(sp_info)
                    log.info(f'歌曲{sp_info.title}，谱面id={sp_info.chart_id}的信息（SP）已获取')
                # 其实先放到all_chart_list然后用两个filter，也行，但是没必要再去写个循环了

    # 把all、ex、sp列表分别写入三张表，注意row是行，col是列
    for sheet, chart_list in [(sheet_ex, ex_chart_list), (sheet_sp, sp_chart_list)]:
        for chart in chart_list:
            for col, key in column_to_write.items():
                sheet.cell(column=col, row=chart.id + 1, value=chart.__getattribute__(key))

    for row, chart in enumerate(all_chart_list, start=2):
        for col, key in column_to_write.items():
            sheet_all.cell(column=col, row=row, value=chart.__getattribute__(key))

    savepath = os.path.join(temp_absdir, 'all_charts.xlsx')
    book.save(savepath)

    log.info(f'乐曲列表已生成完毕，保存于{savepath}')
    log.info('定时任务已处理完成 耗时%.3fs' % (time.time() - start))
